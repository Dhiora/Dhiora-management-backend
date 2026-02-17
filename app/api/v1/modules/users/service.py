import io
from datetime import date
from typing import List, Optional, Tuple
from uuid import UUID

import secrets
from fastapi import UploadFile, status
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.datavalidation import DataValidation

from sqlalchemy import select, func
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.auth.models import Role, StaffProfile, StudentProfile, User, TeacherReferral
from app.auth.referral_code import generate_teacher_referral_code
from app.core.models import AcademicYear, ReferralUsage, SchoolClass, Section, StudentAcademicRecord
from app.auth.security import hash_password
from app.core.exceptions import ServiceError
from app.core.models import Tenant

from app.api.v1.academic_years import service as academic_year_service
from app.api.v1.departments import service as department_service
from app.api.v1.classes import service as class_service
from app.api.v1.sections import service as section_service

from .schemas import (
    DefaultClassPromotion,
    EmployeeCreate,
    EmployeeUpdate,
    EmployeeResponse,
    StaffProfileResponse,
    StudentCreate,
    StudentBulkItem,
    StudentPromote,
    StudentBulkPromote,
    PromotionAction,
    StudentBulkPromoteResult,
    StudentUpdate,
    StudentResponse,
    StudentProfileResponse,
)


DEFAULT_STUDENT_ROLE_NAME = "STUDENT"


async def store_teacher_referral_usage(
    db: AsyncSession,
    tenant_id: UUID,
    referral_code: str,
    student_id: UUID,
    admission_id: UUID,
    academic_year_id: UUID,
) -> None:
    """
    Store that a student was admitted using a teacher's referral code.
    Call only after student and student_academic_record are created, in the same transaction.
    Validates: referral code exists for tenant, teacher is active, no duplicate student/admission.
    Raises ServiceError on validation failure. Does not commit; caller must commit.
    """
    code = (referral_code or "").strip().upper()
    if not code:
        raise ServiceError("Referral code is required when storing referral usage", status.HTTP_400_BAD_REQUEST)

    ref_row = await db.execute(
        select(TeacherReferral).where(
            TeacherReferral.tenant_id == tenant_id,
            TeacherReferral.referral_code == code,
        )
    )
    tr = ref_row.scalar_one_or_none()
    if not tr:
        raise ServiceError(
            "Invalid referral code or code does not belong to this tenant.",
            status.HTTP_400_BAD_REQUEST,
        )

    teacher = await db.get(User, tr.teacher_id)
    if not teacher or teacher.tenant_id != tenant_id:
        raise ServiceError("Referral code references an invalid teacher.", status.HTTP_400_BAD_REQUEST)
    if teacher.status != "ACTIVE":
        raise ServiceError("Referral code belongs to an inactive teacher.", status.HTTP_400_BAD_REQUEST)
    if teacher.role != "Teacher":
        raise ServiceError("Referral code must belong to a teacher.", status.HTTP_400_BAD_REQUEST)

    existing_student = await db.execute(select(ReferralUsage).where(ReferralUsage.student_id == student_id))
    if existing_student.scalar_one_or_none():
        raise ServiceError("This student has already been linked to a referral.", status.HTTP_409_CONFLICT)
    existing_admission = await db.execute(select(ReferralUsage).where(ReferralUsage.admission_id == admission_id))
    if existing_admission.scalar_one_or_none():
        raise ServiceError("This admission is already linked to a referral.", status.HTTP_409_CONFLICT)

    usage = ReferralUsage(
        tenant_id=tenant_id,
        referral_code=code,
        teacher_id=tr.teacher_id,
        student_id=student_id,
        admission_id=admission_id,
        academic_year_id=academic_year_id,
    )
    db.add(usage)


def _to_uuid(value) -> Optional[UUID]:
    if value is None:
        return None
    from uuid import UUID
    return value if isinstance(value, UUID) else UUID(str(value))


async def _get_role_by_id(db: AsyncSession, role_id: UUID, tenant_id: UUID) -> Optional[Role]:
    result = await db.execute(
        select(Role).where(Role.id == role_id, Role.tenant_id == tenant_id)
    )
    return result.scalar_one_or_none()


async def _get_student_role(db: AsyncSession, tenant_id: UUID) -> Optional[Role]:
    """Look up default STUDENT role by name for the tenant."""
    result = await db.execute(
        select(Role).where(Role.tenant_id == tenant_id, Role.name == DEFAULT_STUDENT_ROLE_NAME)
    )
    return result.scalar_one_or_none()


async def _get_role_for_student(db: AsyncSession, tenant_id: UUID, role_id: Optional[UUID]) -> Optional[Role]:
    """Resolve role for a student: if role_id given, use that (must belong to tenant); else default STUDENT role."""
    if role_id is not None:
        return await _get_role_by_id(db, role_id, tenant_id)
    return await _get_student_role(db, tenant_id)


async def _check_duplicate_email(db: AsyncSession, tenant_id: UUID, email: str, exclude_user_id: Optional[UUID] = None) -> bool:
    stmt = select(User).where(User.tenant_id == tenant_id, User.email == email)
    if exclude_user_id is not None:
        stmt = stmt.where(User.id != exclude_user_id)
    result = await db.execute(stmt)
    return result.scalar_one_or_none() is not None


async def _check_duplicate_mobile(db: AsyncSession, tenant_id: UUID, mobile: str, exclude_user_id: Optional[UUID] = None) -> bool:
    if not mobile or not mobile.strip():
        return False
    stmt = select(User).where(User.tenant_id == tenant_id, User.mobile == mobile.strip())
    if exclude_user_id is not None:
        stmt = stmt.where(User.id != exclude_user_id)
    result = await db.execute(stmt)
    return result.scalar_one_or_none() is not None


def _fallback_org_code() -> str:
    """Generate fallback org code for employee_code when tenant has no org_short_code (e.g. ORG7F2)."""
    alphabet = "ABCDEFGHJKLMNPQRSTUVWXYZ23456789"
    return "ORG" + "".join(secrets.choice(alphabet) for _ in range(3))


async def _get_next_employee_sequence(db: AsyncSession, tenant_id: UUID) -> int:
    """Next sequence (1-based) for employee_code per tenant. Zero-padded to 3 digits in format."""
    stmt = (
        select(func.count(StaffProfile.id))
        .select_from(StaffProfile)
        .join(User, StaffProfile.user_id == User.id)
        .where(User.tenant_id == tenant_id)
    )
    result = await db.execute(stmt)
    count = result.scalar() or 0
    return count + 1


async def _get_employee_org_code(db: AsyncSession, tenant_id: UUID) -> str:
    """ORG_CODE for employee_code: tenant.org_short_code if set, else fallback (e.g. ORG7F2)."""
    result = await db.execute(select(Tenant.org_short_code).where(Tenant.id == tenant_id))
    val = result.scalars().first()
    if val is not None and str(val).strip():
        return str(val).strip().upper()[:10]
    return _fallback_org_code()


def _user_to_employee_response(user: User) -> EmployeeResponse:
    sp = user.staff_profile
    return EmployeeResponse(
        id=_to_uuid(user.id),
        tenant_id=_to_uuid(user.tenant_id),
        full_name=user.full_name,
        email=user.email,
        mobile=user.mobile,
        role=user.role,
        role_id=_to_uuid(user.role_id),
        status=user.status,
        user_type=user.user_type,
        created_at=user.created_at,
        staff_profile=StaffProfileResponse(
            id=_to_uuid(sp.id),
            user_id=_to_uuid(sp.user_id),
            employee_code=sp.employee_code,
            department_id=_to_uuid(sp.department_id),
            designation=sp.designation,
            join_date=sp.join_date,
        ) if sp else None,
    )


async def _get_current_student_record(
    db: AsyncSession,
    student_id: UUID,
    tenant_id: UUID,
) -> Optional[StudentAcademicRecord]:
    """Get student_academic_record for current academic year (is_current=true)."""
    result = await db.execute(
        select(StudentAcademicRecord)
        .join(AcademicYear, StudentAcademicRecord.academic_year_id == AcademicYear.id)
        .where(
            StudentAcademicRecord.student_id == student_id,
            AcademicYear.tenant_id == tenant_id,
            AcademicYear.is_current.is_(True),
        )
    )
    return result.scalar_one_or_none()


async def _user_to_student_response(
    db: AsyncSession,
    user: User,
    tenant_id: UUID,
) -> StudentResponse:
    """Build StudentResponse. class_id/section_id from current student_academic_record; class_name/section_name from core.classes and core.sections."""
    sp = user.student_profile
    class_id, section_id, roll_number = None, None, (sp.roll_number if sp else None)
    class_name, section_name = None, None
    if sp and user.id:
        rec = await _get_current_student_record(db, user.id, tenant_id)
        if rec:
            class_id, section_id, roll_number = rec.class_id, rec.section_id, rec.roll_number or (sp.roll_number)
            if rec.class_id:
                school_class = await db.get(SchoolClass, rec.class_id)
                class_name = school_class.name if school_class else None
            if rec.section_id:
                section = await db.get(Section, rec.section_id)
                section_name = section.name if section else None
    return StudentResponse(
        id=_to_uuid(user.id),
        tenant_id=_to_uuid(user.tenant_id),
        full_name=user.full_name,
        email=user.email,
        mobile=user.mobile,
        role=user.role,
        role_id=_to_uuid(user.role_id),
        status=user.status,
        user_type=user.user_type,
        created_at=user.created_at,
        student_profile=StudentProfileResponse(
            id=_to_uuid(sp.id),
            user_id=_to_uuid(sp.user_id),
            roll_number=roll_number,
            class_id=_to_uuid(class_id),
            section_id=_to_uuid(section_id),
            class_name=class_name,
            section_name=section_name,
        ) if sp else None,
    )


# ----- Employee -----
async def create_employee(
    db: AsyncSession,
    tenant_id: UUID,
    payload: EmployeeCreate,
) -> EmployeeResponse:
    """Create employee: validate role_id and department_id, auto-generate employee_code, create user + staff_profile. Raises ServiceError 400/409."""
    role = await _get_role_by_id(db, payload.role_id, tenant_id)
    if not role:
        raise ServiceError("role_id does not belong to this tenant", status.HTTP_400_BAD_REQUEST)

    dept = await department_service.get_department_by_id_for_tenant(db, tenant_id, payload.department_id, active_only=True)
    if not dept:
        raise ServiceError("Invalid department", status.HTTP_400_BAD_REQUEST)

    if await _check_duplicate_email(db, tenant_id, payload.email):
        raise ServiceError("Email already exists for this tenant", status.HTTP_409_CONFLICT)
    if payload.mobile and await _check_duplicate_mobile(db, tenant_id, payload.mobile):
        raise ServiceError("Mobile number already exists for this tenant", status.HTTP_409_CONFLICT)

    org_code = await _get_employee_org_code(db, tenant_id)
    next_seq = await _get_next_employee_sequence(db, tenant_id)
    employee_code = f"{org_code}-EMP-{next_seq:03d}"

    try:
        password_hash = hash_password(payload.password)
        user = User(
            tenant_id=tenant_id,
            full_name=payload.full_name,
            email=payload.email,
            mobile=payload.mobile,
            password_hash=password_hash,
            role=role.name,
            role_id=role.id,
            status="ACTIVE",
            source="EMPLOYEE",
            user_type="employee",
        )
        db.add(user)
        await db.flush()

        staff_profile = StaffProfile(
            user_id=user.id,
            employee_code=employee_code,
            department_id=payload.department_id,
            designation=payload.designation,
            join_date=payload.join_date,
        )
        db.add(staff_profile)
        await db.commit()
        await db.refresh(user)
        await db.refresh(staff_profile)

        if role.name == "Teacher":
            max_attempts = 5
            for attempt in range(max_attempts):
                referral_code = generate_teacher_referral_code(payload.full_name)
                ref = TeacherReferral(
                    teacher_id=user.id,
                    tenant_id=tenant_id,
                    referral_code=referral_code,
                )
                db.add(ref)
                try:
                    await db.commit()
                    break
                except IntegrityError:
                    await db.rollback()
                    if attempt == max_attempts - 1:
                        raise ServiceError(
                            "Could not generate unique referral code for teacher after retries",
                            status.HTTP_500_INTERNAL_SERVER_ERROR,
                        )

        user.staff_profile = staff_profile
        return _user_to_employee_response(user)
    except IntegrityError:
        await db.rollback()
        raise ServiceError("Duplicate email or mobile for this tenant", status.HTTP_409_CONFLICT)


async def list_employees(
    db: AsyncSession,
    tenant_id: UUID,
) -> List[EmployeeResponse]:
    result = await db.execute(
        select(User)
        .where(User.tenant_id == tenant_id, User.user_type == "employee")
        .options(selectinload(User.staff_profile))
        .order_by(User.full_name)
    )
    users = result.scalars().all()
    return [_user_to_employee_response(u) for u in users]


async def get_employee(db: AsyncSession, tenant_id: UUID, user_id: UUID) -> Optional[EmployeeResponse]:
    result = await db.execute(
        select(User)
        .where(User.id == user_id, User.tenant_id == tenant_id, User.user_type == "employee")
        .options(selectinload(User.staff_profile))
    )
    user = result.scalar_one_or_none()
    return _user_to_employee_response(user) if user else None


async def update_employee(
    db: AsyncSession,
    tenant_id: UUID,
    user_id: UUID,
    payload: EmployeeUpdate,
) -> Optional[EmployeeResponse]:
    result = await db.execute(
        select(User)
        .where(User.id == user_id, User.tenant_id == tenant_id, User.user_type == "employee")
        .options(selectinload(User.staff_profile))
    )
    user = result.scalar_one_or_none()
    if not user:
        return None

    if payload.email is not None and await _check_duplicate_email(db, tenant_id, payload.email, exclude_user_id=user_id):
        raise ServiceError("Email already exists for this tenant", status.HTTP_409_CONFLICT)
    if payload.mobile is not None and payload.mobile.strip() and await _check_duplicate_mobile(db, tenant_id, payload.mobile, exclude_user_id=user_id):
        raise ServiceError("Mobile number already exists for this tenant", status.HTTP_409_CONFLICT)

    if payload.role_id is not None:
        role = await _get_role_by_id(db, payload.role_id, tenant_id)
        if not role:
            raise ServiceError("role_id does not belong to this tenant", status.HTTP_400_BAD_REQUEST)
        user.role_id = role.id
        user.role = role.name

    if payload.full_name is not None:
        user.full_name = payload.full_name
    if payload.email is not None:
        user.email = payload.email
    if payload.mobile is not None:
        user.mobile = payload.mobile or None
    if payload.password is not None:
        user.password_hash = hash_password(payload.password)
    if payload.status is not None:
        user.status = payload.status

    if user.staff_profile:
        # employee_code is NOT editable after creation
        if payload.department_id is not None:
            dept = await department_service.get_department_by_id_for_tenant(db, tenant_id, payload.department_id, active_only=True)
            if not dept:
                raise ServiceError("Invalid department", status.HTTP_400_BAD_REQUEST)
            user.staff_profile.department_id = payload.department_id
        if payload.designation is not None:
            user.staff_profile.designation = payload.designation
        if payload.join_date is not None:
            user.staff_profile.join_date = payload.join_date

    try:
        await db.commit()
        await db.refresh(user)
        if user.staff_profile:
            await db.refresh(user.staff_profile)
        return _user_to_employee_response(user)
    except IntegrityError:
        await db.rollback()
        raise ServiceError("Duplicate email or mobile for this tenant", status.HTTP_409_CONFLICT)


async def delete_employee(db: AsyncSession, tenant_id: UUID, user_id: UUID) -> bool:
    result = await db.execute(
        select(User).where(User.id == user_id, User.tenant_id == tenant_id, User.user_type == "employee")
    )
    user = result.scalar_one_or_none()
    if not user:
        return False
    await db.delete(user)
    await db.commit()
    return True


# ----- Student -----
async def create_student(
    db: AsyncSession,
    tenant_id: UUID,
    payload: StudentCreate,
) -> StudentResponse:
    """Create student: requires admission-open academic year. Creates user + student_profile + student_academic_records."""
    ay = await academic_year_service.get_admission_open_academic_year(db, tenant_id)
    if not ay:
        raise ServiceError(
            "No academic year is open for admissions. Create an academic year with is_current=true, status=ACTIVE, "
            "and admissions_allowed=true.",
            status.HTTP_400_BAD_REQUEST,
        )

    role = await _get_role_for_student(db, tenant_id, getattr(payload, "role_id", None))
    if not role:
        raise ServiceError(
            "Role not found. Provide a valid role_id for this tenant or create a role named 'STUDENT'.",
            status.HTTP_400_BAD_REQUEST,
        )

    school_class = await class_service.get_class_by_id_for_tenant(db, tenant_id, payload.class_id, active_only=True)
    if not school_class:
        raise ServiceError("Invalid class", status.HTTP_400_BAD_REQUEST)
    section = await section_service.get_section_by_id_for_tenant(db, tenant_id, payload.section_id, active_only=True)
    if not section:
        raise ServiceError("Invalid section", status.HTTP_400_BAD_REQUEST)
    if section.class_id != payload.class_id:
        raise ServiceError("Section does not belong to the selected class", status.HTTP_400_BAD_REQUEST)

    if await _check_duplicate_email(db, tenant_id, payload.email):
        raise ServiceError("Email already exists for this tenant", status.HTTP_409_CONFLICT)
    if payload.mobile and await _check_duplicate_mobile(db, tenant_id, payload.mobile):
        raise ServiceError("Mobile number already exists for this tenant", status.HTTP_409_CONFLICT)

    try:
        password_hash = hash_password(payload.password)
        user = User(
            tenant_id=tenant_id,
            full_name=payload.full_name,
            email=payload.email,
            mobile=payload.mobile,
            password_hash=password_hash,
            role=role.name,
            role_id=role.id,
            status="ACTIVE",
            source="STUDENT",
            user_type="student",
        )
        db.add(user)
        await db.flush()

        student_profile = StudentProfile(
            user_id=user.id,
            roll_number=payload.roll_number,
        )
        db.add(student_profile)
        await db.flush()

        record = StudentAcademicRecord(
            student_id=user.id,
            academic_year_id=ay.id,
            class_id=payload.class_id,
            section_id=payload.section_id,
            roll_number=payload.roll_number,
            status="ACTIVE",
        )
        db.add(record)
        await db.flush()

        if getattr(payload, "referral_code", None) and str(payload.referral_code).strip():
            await store_teacher_referral_usage(
                db,
                tenant_id,
                payload.referral_code,
                user.id,
                record.id,
                ay.id,
            )

        await db.commit()
        await db.refresh(user)
        await db.refresh(student_profile)
        user.student_profile = student_profile
        return await _user_to_student_response(db, user, tenant_id)
    except IntegrityError:
        await db.rollback()
        raise ServiceError("Duplicate email or mobile for this tenant", status.HTTP_409_CONFLICT)


EXCEL_MAX_ROWS = 500
EXCEL_REQUIRED_HEADERS = ("full_name", "email", "mobile", "password", "roll_number", "class_id", "section_id")
TEMPLATE_HEADERS = ("full_name", "email", "mobile", "password", "roll_number", "class_name", "section_display")
STUDENTS_SHEET_NAME = "Students"
CLASSES_SHEET_NAME = "Classes"
SECTIONS_SHEET_NAME = "Sections"


async def build_student_upload_template(
    db: AsyncSession,
    tenant_id: UUID,
) -> bytes:
    """Build Excel template with Classes and Sections sheets and dropdowns on Students sheet."""
    wb = Workbook()
    classes_list = await class_service.list_classes(db, tenant_id, active_only=True)
    class_name_to_id = {c.name: str(c.id) for c in classes_list}

    # Sheet: Classes (class_name, class_id)
    ws_classes = wb.active
    ws_classes.title = CLASSES_SHEET_NAME
    ws_classes.append(["class_name", "class_id"])
    for c in classes_list:
        ws_classes.append([c.name, str(c.id)])

    # Sheet: Sections (section_display, section_id, class_id)
    ws_sections = wb.create_sheet(SECTIONS_SHEET_NAME)
    ws_sections.append(["section_display", "section_id", "class_id"])
    section_displays: List[str] = []
    for c in classes_list:
        sections = await section_service.list_sections(db, tenant_id, active_only=True, class_id=c.id)
        for s in sections:
            display = f"{s.name} ({c.name})"
            section_displays.append(display)
            ws_sections.append([display, str(s.id), str(c.id)])

    # Sheet: Students (main data entry with dropdowns)
    ws_students = wb.create_sheet(STUDENTS_SHEET_NAME, 0)
    ws_students.append(list(TEMPLATE_HEADERS))

    n_classes = len(classes_list)
    n_sections = len(section_displays)
    if n_classes > 0:
        dv_class = DataValidation(
            type="list",
            formula1=f"'{CLASSES_SHEET_NAME}'!$A$2:$A${1 + n_classes}",
            allow_blank=False,
        )
        dv_class.error = "Select a value from the Class dropdown"
        ws_students.add_data_validation(dv_class)
        dv_class.add(f"F2:F{EXCEL_MAX_ROWS + 1}")
    if n_sections > 0:
        dv_section = DataValidation(
            type="list",
            formula1=f"'{SECTIONS_SHEET_NAME}'!$A$2:$A${1 + n_sections}",
            allow_blank=False,
        )
        dv_section.error = "Select a value from the Section dropdown"
        ws_students.add_data_validation(dv_section)
        dv_section.add(f"G2:G{EXCEL_MAX_ROWS + 1}")

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


async def parse_students_excel(
    file: UploadFile,
    db: Optional[AsyncSession] = None,
    tenant_id: Optional[UUID] = None,
) -> List[StudentBulkItem]:
    """Parse uploaded Excel file into list of StudentBulkItem. First row = headers. Max 500 data rows. Raises ValueError on invalid format."""
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise ValueError("File must be an Excel file (.xlsx)")

    content = await file.read()
    if not content:
        raise ValueError("File is empty")

    try:
        wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        raise ValueError(f"Invalid Excel file: {e}") from e

    ws = wb.active
    if not ws:
        raise ValueError("Excel file has no active sheet")

    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if not header_row:
        raise ValueError("Excel file has no header row")

    def _norm(s: str) -> str:
        return (str(s).strip().lower() if s is not None else "").replace(" ", "_")

    header_row = [_norm(c) for c in header_row]
    use_template_format = "class_name" in header_row and "section_display" in header_row
    use_id_format = "class_id" in header_row and "section_id" in header_row

    if use_template_format:
        required = list(TEMPLATE_HEADERS)
        if db is None or tenant_id is None:
            raise ValueError("Template format (class_name, section_display) requires db and tenant_id to resolve")
    elif use_id_format:
        required = list(EXCEL_REQUIRED_HEADERS)
    else:
        raise ValueError(
            f"Excel must have either (class_id, section_id) or (class_name, section_display). Found: {header_row}"
        )

    col_idx = {}
    for h in required:
        try:
            col_idx[h] = header_row.index(h)
        except ValueError:
            raise ValueError(f"Missing required column: {h}. Found: {header_row}")

    items: List[StudentBulkItem] = []
    for row_num, row in enumerate(rows_iter, start=2):
        if row_num - 1 > EXCEL_MAX_ROWS:
            raise ValueError(f"Maximum {EXCEL_MAX_ROWS} data rows allowed")
        if not row or all(c is None or (isinstance(c, str) and not c.strip()) for c in row):
            continue
        try:
            full_name = _cell_str(row, col_idx["full_name"])
            email = _cell_str(row, col_idx["email"])
            mobile = _cell_str(row, col_idx["mobile"]) or None
            password = _cell_str(row, col_idx["password"])
            roll_number = _cell_str(row, col_idx["roll_number"]) or None
            if not full_name or not email or not password:
                raise ValueError(f"Row {row_num}: full_name, email, password are required")
            if use_template_format:
                class_name_val = _cell_str(row, col_idx["class_name"])
                section_display_val = _cell_str(row, col_idx["section_display"])
                if not class_name_val or not section_display_val:
                    raise ValueError(f"Row {row_num}: class_name and section_display are required")
                class_id, section_id = await _resolve_class_section_from_template(
                    db, tenant_id, class_name_val.strip(), section_display_val.strip()
                )
            else:
                class_id_str = _cell_str(row, col_idx["class_id"])
                section_id_str = _cell_str(row, col_idx["section_id"])
                if not class_id_str or not section_id_str:
                    raise ValueError(f"Row {row_num}: class_id and section_id are required")
                class_id = UUID(str(class_id_str).strip())
                section_id = UUID(str(section_id_str).strip())
            items.append(
                StudentBulkItem(
                    full_name=full_name,
                    email=email,
                    mobile=mobile,
                    password=password,
                    roll_number=roll_number,
                    class_id=class_id,
                    section_id=section_id,
                )
            )
        except (ValueError, TypeError) as e:
            raise ValueError(f"Row {row_num}: {e}") from e
    wb.close()
    return items


async def parse_students_excel_with_errors(
    file: UploadFile,
    db: AsyncSession,
    tenant_id: UUID,
) -> Tuple[List[Tuple[StudentBulkItem, dict]], List[Tuple[dict, str]]]:
    """
    Parse Excel row by row. Returns (success_list, failed_list).
    success_list = [(StudentBulkItem, row_dict), ...]
    failed_list = [(row_dict, reason), ...] for rows that failed to parse.
    """
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise ValueError("File must be an Excel file (.xlsx)")
    content = await file.read()
    if not content:
        raise ValueError("File is empty")
    try:
        wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        raise ValueError(f"Invalid Excel file: {e}") from e
    ws = wb.active
    if not ws:
        raise ValueError("Excel file has no active sheet")
    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if not header_row:
        raise ValueError("Excel file has no header row")

    def _norm(s: str) -> str:
        return (str(s).strip().lower() if s is not None else "").replace(" ", "_")

    header_row_norm = [_norm(c) for c in header_row]
    use_template_format = "class_name" in header_row_norm and "section_display" in header_row_norm
    use_id_format = "class_id" in header_row_norm and "section_id" in header_row_norm
    if use_template_format:
        required = list(TEMPLATE_HEADERS)
    elif use_id_format:
        required = list(EXCEL_REQUIRED_HEADERS)
    else:
        raise ValueError(
            f"Excel must have either (class_id, section_id) or (class_name, section_display). Found: {header_row_norm}"
        )
    col_idx = {}
    for h in required:
        try:
            col_idx[h] = header_row_norm.index(h)
        except ValueError:
            wb.close()
            raise ValueError(f"Missing required column: {h}. Found: {header_row_norm}")

    success: List[Tuple[StudentBulkItem, dict]] = []
    failed: List[Tuple[dict, str]] = []
    headers_for_dict = [header_row[i] if i < len(header_row) else f"col_{i}" for i in range(max(col_idx.values()) + 1)]

    for row_num, row in enumerate(rows_iter, start=2):
        if row_num - 1 > EXCEL_MAX_ROWS:
            break
        if not row or all(c is None or (isinstance(c, str) and not c.strip()) for c in row):
            continue
        row_dict = {}
        for i, h in enumerate(header_row_norm):
            if i < len(row):
                row_dict[h] = row[i] if row[i] is None else str(row[i]).strip()
            else:
                row_dict[h] = ""
        try:
            full_name = _cell_str(row, col_idx["full_name"])
            email = _cell_str(row, col_idx["email"])
            mobile = _cell_str(row, col_idx["mobile"]) or None
            password = _cell_str(row, col_idx["password"])
            roll_number = _cell_str(row, col_idx["roll_number"]) or None
            if not full_name or not email or not password:
                failed.append((row_dict, "full_name, email, password are required"))
                continue
            if use_template_format:
                class_name_val = _cell_str(row, col_idx["class_name"])
                section_display_val = _cell_str(row, col_idx["section_display"])
                if not class_name_val or not section_display_val:
                    failed.append((row_dict, "class_name and section_display are required"))
                    continue
                class_id, section_id = await _resolve_class_section_from_template(
                    db, tenant_id, class_name_val.strip(), section_display_val.strip()
                )
            else:
                class_id_str = _cell_str(row, col_idx["class_id"])
                section_id_str = _cell_str(row, col_idx["section_id"])
                if not class_id_str or not section_id_str:
                    failed.append((row_dict, "class_id and section_id are required"))
                    continue
                try:
                    class_id = UUID(str(class_id_str).strip())
                    section_id = UUID(str(section_id_str).strip())
                except (ValueError, TypeError) as e:
                    failed.append((row_dict, f"Invalid UUID for class_id or section_id: {e}"))
                    continue
            item = StudentBulkItem(
                full_name=full_name,
                email=email,
                mobile=mobile,
                password=password,
                roll_number=roll_number,
                class_id=class_id,
                section_id=section_id,
            )
            success.append((item, row_dict))
        except (ValueError, TypeError) as e:
            failed.append((row_dict, str(e)))
    wb.close()
    return success, failed


async def _resolve_class_section_from_template(
    db: AsyncSession,
    tenant_id: UUID,
    class_name: str,
    section_display: str,
) -> Tuple[UUID, UUID]:
    """Parse section_display 'SectionName (ClassName)' and resolve to (class_id, section_id)."""
    if " (" not in section_display or not section_display.strip().endswith(")"):
        raise ValueError(f"section_display must be 'Section Name (Class Name)': got '{section_display}'")
    part = section_display.rsplit(" (", 1)
    section_name = part[0].strip()
    class_name_in_display = part[1].rstrip(")").strip()
    if not section_name or not class_name_in_display:
        raise ValueError(f"section_display must be 'Section Name (Class Name)': got '{section_display}'")
    classes_list = await class_service.list_classes(db, tenant_id, active_only=True)
    class_by_name = {c.name: c.id for c in classes_list}
    if class_name not in class_by_name:
        raise ValueError(f"Class not found: '{class_name}'")
    class_id = class_by_name[class_name]
    if class_name_in_display != class_name:
        raise ValueError(f"Section's class '{class_name_in_display}' does not match selected class '{class_name}'")
    sections_list = await section_service.list_sections(db, tenant_id, active_only=True, class_id=class_id)
    section_by_name = {s.name: s.id for s in sections_list}
    if section_name not in section_by_name:
        raise ValueError(f"Section not found: '{section_name}' in class '{class_name}'")
    return class_id, section_by_name[section_name]


def _cell_str(row: tuple, col: int) -> str:
    if col >= len(row):
        return ""
    v = row[col]
    if v is None:
        return ""
    return str(v).strip()


async def create_students_bulk(
    db: AsyncSession,
    tenant_id: UUID,
    items: List[StudentBulkItem],
) -> List[StudentResponse]:
    """Create multiple students in one transaction. Validates class/section, deduplicates email/mobile. All-or-nothing."""
    created, _ = await create_students_bulk_with_failures(db, tenant_id, [(item, None) for item in items])
    return created


def _build_error_excel(failed_rows: List[Tuple[dict, str]], headers: Optional[List[str]] = None) -> bytes:
    """Build Excel file with failed rows and reason column."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Upload errors"
    if not failed_rows:
        ws.append(["No failed rows"])
    else:
        row_dict, _ = failed_rows[0]
        col_headers = headers or list(row_dict.keys())
        ws.append(col_headers + ["reason"])
        for row_dict, reason in failed_rows:
            row = [row_dict.get(h, "") for h in col_headers]
            ws.append(row + [reason])
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


async def create_students_bulk_with_failures(
    db: AsyncSession,
    tenant_id: UUID,
    items_with_row_data: List[Tuple[StudentBulkItem, Optional[dict]]],
) -> Tuple[List[StudentResponse], List[Tuple[dict, str]]]:
    """
    Validate each item, create valid ones in one transaction, return (created, failed).
    Requires admission-open academic year. Each item can have its own role_id; if omitted, default STUDENT role is used.
    """
    ay = await academic_year_service.get_admission_open_academic_year(db, tenant_id)
    if not ay:
        msg = "No academic year is open for admissions. Create an academic year with is_current=true, status=ACTIVE, and admissions_allowed=true."
        return [], [(_item_to_row_dict(items_with_row_data[0][0]) if items_with_row_data else {}, msg)]

    failed: List[Tuple[dict, str]] = []
    to_create: List[Tuple[StudentBulkItem, Optional[dict], Role]] = []

    emails_seen: set = set()
    mobiles_seen: set = set()

    for item, row_dict in items_with_row_data:
        rd = row_dict or _item_to_row_dict(item)
        role = await _get_role_for_student(db, tenant_id, getattr(item, "role_id", None))
        if not role:
            failed.append((rd, "Role not found. Provide a valid role_id for this tenant or create a role named 'STUDENT'."))
            continue
        err = None
        if item.email.lower() in emails_seen:
            err = f"Duplicate email in upload: {item.email}"
        elif item.mobile and item.mobile.strip():
            m = item.mobile.strip()
            if m in mobiles_seen:
                err = f"Duplicate mobile in upload: {item.mobile}"
        if err:
            failed.append((rd, err))
            continue
        emails_seen.add(item.email.lower())
        if item.mobile and item.mobile.strip():
            mobiles_seen.add(item.mobile.strip())

        school_class = await class_service.get_class_by_id_for_tenant(db, tenant_id, item.class_id, active_only=True)
        if not school_class:
            failed.append((rd, f"Invalid or inactive class for this tenant (class_id={item.class_id})"))
            continue
        section = await section_service.get_section_by_id_for_tenant(db, tenant_id, item.section_id, active_only=True)
        if not section:
            failed.append((rd, f"Invalid or inactive section for this tenant (section_id={item.section_id})"))
            continue
        if getattr(section, "class_id", None) != item.class_id:
            failed.append((rd, "Section does not belong to the given class"))
            continue
        if await _check_duplicate_email(db, tenant_id, item.email):
            failed.append((rd, f"Email already exists for this tenant: {item.email}"))
            continue
        if item.mobile and await _check_duplicate_mobile(db, tenant_id, item.mobile):
            failed.append((rd, f"Mobile number already exists for this tenant: {item.mobile}"))
            continue
        to_create.append((item, row_dict, role))

    created_users: List[User] = []
    try:
        for item, _, role in to_create:
            password_hash = hash_password(item.password)
            user = User(
                tenant_id=tenant_id,
                full_name=item.full_name,
                email=item.email,
                mobile=item.mobile,
                password_hash=password_hash,
                role=role.name,
                role_id=role.id,
                status="ACTIVE",
                source="STUDENT",
                user_type="student",
            )
            db.add(user)
            await db.flush()
            student_profile = StudentProfile(
                user_id=user.id,
                roll_number=item.roll_number,
            )
            db.add(student_profile)
            await db.flush()
            record = StudentAcademicRecord(
                student_id=user.id,
                academic_year_id=ay.id,
                class_id=item.class_id,
                section_id=item.section_id,
                roll_number=item.roll_number,
                status="ACTIVE",
            )
            db.add(record)
            user.student_profile = student_profile
            created_users.append(user)
        await db.commit()
    except IntegrityError:
        await db.rollback()
        for item, row_dict, _ in to_create:
            rd = row_dict or _item_to_row_dict(item)
            failed.append((rd, "Duplicate email or mobile for this tenant (database constraint)"))
        return [], failed

    # Re-load users with student_profile eagerly (avoid lazy load in async)
    user_ids = [u.id for u in created_users]
    result = await db.execute(
        select(User).where(User.id.in_(user_ids)).options(selectinload(User.student_profile))
    )
    users_by_id = {u.id: u for u in result.scalars().unique().all()}
    ordered_users = [users_by_id[uid] for uid in user_ids]
    responses = []
    for u in ordered_users:
        responses.append(await _user_to_student_response(db, u, tenant_id))
    return responses, failed


def _item_to_row_dict(item: StudentBulkItem) -> dict:
    """Minimal row dict from StudentBulkItem for error report."""
    d = {
        "full_name": item.full_name,
        "email": item.email,
        "mobile": item.mobile or "",
        "password": "(hidden)",
        "roll_number": item.roll_number or "",
        "class_id": str(item.class_id),
        "section_id": str(item.section_id),
    }
    if getattr(item, "role_id", None) is not None:
        d["role_id"] = str(item.role_id)
    return d


async def list_students(
    db: AsyncSession,
    tenant_id: UUID,
) -> List[StudentResponse]:
    result = await db.execute(
        select(User)
        .where(User.tenant_id == tenant_id, User.user_type == "student")
        .options(selectinload(User.student_profile))
        .order_by(User.full_name)
    )
    users = result.scalars().all()
    return [await _user_to_student_response(db, u, tenant_id) for u in users]


async def get_student(db: AsyncSession, tenant_id: UUID, user_id: UUID) -> Optional[StudentResponse]:
    result = await db.execute(
        select(User)
        .where(User.id == user_id, User.tenant_id == tenant_id, User.user_type == "student")
        .options(selectinload(User.student_profile))
    )
    user = result.scalar_one_or_none()
    return await _user_to_student_response(db, user, tenant_id) if user else None


async def update_student(
    db: AsyncSession,
    tenant_id: UUID,
    user_id: UUID,
    payload: StudentUpdate,
) -> Optional[StudentResponse]:
    result = await db.execute(
        select(User)
        .where(User.id == user_id, User.tenant_id == tenant_id, User.user_type == "student")
        .options(selectinload(User.student_profile))
    )
    user = result.scalar_one_or_none()
    if not user:
        return None

    if payload.email is not None and await _check_duplicate_email(db, tenant_id, payload.email, exclude_user_id=user_id):
        raise ServiceError("Email already exists for this tenant", status.HTTP_409_CONFLICT)
    if payload.mobile is not None and payload.mobile.strip() and await _check_duplicate_mobile(db, tenant_id, payload.mobile, exclude_user_id=user_id):
        raise ServiceError("Mobile number already exists for this tenant", status.HTTP_409_CONFLICT)

    if payload.full_name is not None:
        user.full_name = payload.full_name
    if payload.email is not None:
        user.email = payload.email
    if payload.mobile is not None:
        user.mobile = payload.mobile or None
    if payload.password is not None:
        user.password_hash = hash_password(payload.password)
    if payload.status is not None:
        user.status = payload.status
    if payload.role_id is not None:
        role = await _get_role_by_id(db, payload.role_id, tenant_id)
        if not role:
            raise ServiceError("Role not found for this tenant", status.HTTP_400_BAD_REQUEST)
        user.role_id = role.id
        user.role = role.name

    if user.student_profile:
        if payload.roll_number is not None:
            user.student_profile.roll_number = payload.roll_number
    # Update current student_academic_record for class/section (not student_profile)
    if payload.class_id is not None or payload.section_id is not None:
        rec = await _get_current_student_record(db, user.id, tenant_id)
        if rec:
            if payload.class_id is not None:
                school_class = await class_service.get_class_by_id_for_tenant(db, tenant_id, payload.class_id, active_only=True)
                if not school_class:
                    raise ServiceError("Invalid class", status.HTTP_400_BAD_REQUEST)
                rec.class_id = payload.class_id
            if payload.section_id is not None:
                section = await section_service.get_section_by_id_for_tenant(db, tenant_id, payload.section_id, active_only=True)
                if not section:
                    raise ServiceError("Invalid section", status.HTTP_400_BAD_REQUEST)
                if rec.class_id != section.class_id:
                    raise ServiceError("Section does not belong to the selected class", status.HTTP_400_BAD_REQUEST)
                rec.section_id = payload.section_id
        else:
            raise ServiceError("No current academic record for this student", status.HTTP_400_BAD_REQUEST)

    try:
        await db.commit()
        await db.refresh(user)
        if user.student_profile:
            await db.refresh(user.student_profile)
        return await _user_to_student_response(db, user, tenant_id)
    except IntegrityError:
        await db.rollback()
        raise ServiceError("Duplicate email or mobile for this tenant", status.HTTP_409_CONFLICT)


async def promote_student(
    db: AsyncSession,
    tenant_id: UUID,
    student_id: UUID,
    payload: StudentPromote,
) -> StudentResponse:
    """
    Promote student to new academic year. Old record status → PROMOTED; creates NEW student_academic_record.
    Does NOT update existing records; preserves full academic history.
    """
    ay = await academic_year_service.get_academic_year(db, tenant_id, payload.academic_year_id)
    if not ay:
        raise ServiceError("Academic year not found", status.HTTP_404_NOT_FOUND)
    if ay.status != "ACTIVE":
        raise ServiceError("Cannot promote to a CLOSED academic year", status.HTTP_400_BAD_REQUEST)
    school_class = await class_service.get_class_by_id_for_tenant(db, tenant_id, payload.class_id, active_only=True)
    if not school_class:
        raise ServiceError("Invalid class", status.HTTP_400_BAD_REQUEST)
    section = await section_service.get_section_by_id_for_tenant(db, tenant_id, payload.section_id, active_only=True)
    if not section:
        raise ServiceError("Invalid section", status.HTTP_400_BAD_REQUEST)
    if section.class_id != payload.class_id:
        raise ServiceError("Section does not belong to the selected class", status.HTTP_400_BAD_REQUEST)
    result = await db.execute(
        select(User).where(
            User.id == student_id,
            User.tenant_id == tenant_id,
            User.user_type == "student",
        ).options(selectinload(User.student_profile))
    )
    user = result.scalar_one_or_none()
    if not user:
        raise ServiceError("Student not found", status.HTTP_404_NOT_FOUND)
    # Check if record already exists for new year
    existing = await db.execute(
        select(StudentAcademicRecord).where(
            StudentAcademicRecord.student_id == student_id,
            StudentAcademicRecord.academic_year_id == payload.academic_year_id,
        )
    )
    if existing.scalar_one_or_none():
        raise ServiceError(
            f"Student already has an academic record for academic year {payload.academic_year_id}",
            status.HTTP_409_CONFLICT,
        )
    # Set current record status to PROMOTED
    current_rec = await _get_current_student_record(db, student_id, tenant_id)
    if current_rec:
        current_rec.status = "PROMOTED"
    # Create new record
    new_rec = StudentAcademicRecord(
        student_id=student_id,
        academic_year_id=payload.academic_year_id,
        class_id=payload.class_id,
        section_id=payload.section_id,
        roll_number=payload.roll_number,
        status="ACTIVE",
    )
    db.add(new_rec)
    await db.commit()
    await db.refresh(user)
    return await _user_to_student_response(db, user, tenant_id)


async def _resolve_section_for_promotion(
    db: AsyncSession,
    tenant_id: UUID,
    to_class_id: UUID,
    section_behavior: str,
    current_section_name: Optional[str],
    current_section_id: UUID,
) -> Tuple[Optional[UUID], Optional[str]]:
    """Resolve target section_id. Returns (section_id, error_reason)."""
    sections_list = await section_service.list_sections(db, tenant_id, active_only=True, class_id=to_class_id)
    if not sections_list:
        return None, f"No sections in target class {to_class_id}"
    if section_behavior == "AUTO":
        return sections_list[0].id, None
    if section_behavior == "SAME" and current_section_name:
        for s in sections_list:
            if s.name == current_section_name:
                return s.id, None
        return None, f"No section '{current_section_name}' in target class"
    if section_behavior == "MANUAL":
        return None, "MANUAL section behavior requires student_overrides"
    return None, "Unknown section behavior"


async def promote_students_bulk(
    db: AsyncSession,
    tenant_id: UUID,
    payload: StudentBulkPromote,
    preview: bool = False,
) -> StudentBulkPromoteResult:
    """
    Promote students from source to target academic year.
    RETAIN = same class/section. PROMOTE = new class/section.
    default_section_behavior: AUTO | SAME | MANUAL.
    Transactional; preview=true returns actions without committing.
    """
    # Validate source academic year
    source_ay = await academic_year_service.get_academic_year(db, tenant_id, payload.source_academic_year_id)
    if not source_ay:
        raise ServiceError("Source academic year not found", status.HTTP_404_NOT_FOUND)

    # Validate target academic year
    target_ay = await academic_year_service.get_academic_year(db, tenant_id, payload.target_academic_year_id)
    if not target_ay:
        raise ServiceError("Target academic year not found", status.HTTP_404_NOT_FOUND)
    if target_ay.status != "ACTIVE":
        raise ServiceError("Target academic year must be ACTIVE", status.HTTP_400_BAD_REQUEST)
    if not getattr(target_ay, "admissions_allowed", True):
        raise ServiceError("Target academic year must have admissions_allowed=true", status.HTTP_400_BAD_REQUEST)

    # Validate default_section_behavior
    sec_behavior = (payload.default_section_behavior or "AUTO").upper()
    if sec_behavior not in ("AUTO", "SAME", "MANUAL"):
        raise ServiceError(
            "default_section_behavior must be AUTO, SAME, or MANUAL",
            status.HTTP_400_BAD_REQUEST,
        )

    # Build default class mapping: from_class_id -> to_class_id
    default_class_map = {str(m.from_class_id): m.to_class_id for m in payload.default_class_promotion}

    # Build override lookup: student_id -> (action, to_class_id?, to_section_id?)
    override_map = {}
    for o in payload.student_overrides:
        if o.action not in ("RETAIN", "PROMOTE"):
            raise ServiceError(f"Invalid override action: {o.action}", status.HTTP_400_BAD_REQUEST)
        if o.action == "PROMOTE" and not o.to_class_id:
            raise ServiceError(f"PROMOTE override requires to_class_id for student {o.student_id}", status.HTTP_400_BAD_REQUEST)
        override_map[str(o.student_id)] = (o.action, o.to_class_id, o.to_section_id)

    # Validate PROMOTE overrides: to_class, to_section exist
    for o in payload.student_overrides:
        if o.action == "PROMOTE" and o.to_class_id:
            school_class = await class_service.get_class_by_id_for_tenant(db, tenant_id, o.to_class_id, active_only=True)
            if not school_class:
                raise ServiceError(f"Invalid to_class_id in override: {o.to_class_id}", status.HTTP_400_BAD_REQUEST)
            if o.to_section_id:
                section = await section_service.get_section_by_id_for_tenant(db, tenant_id, o.to_section_id, active_only=True)
                if not section or section.class_id != o.to_class_id:
                    raise ServiceError(f"Invalid to_section_id in override: {o.to_section_id}", status.HTTP_400_BAD_REQUEST)

    # Validate default_class_promotion: to_class exists
    for m in payload.default_class_promotion:
        school_class = await class_service.get_class_by_id_for_tenant(db, tenant_id, m.to_class_id, active_only=True)
        if not school_class:
            raise ServiceError(f"Invalid to_class_id in default_class_promotion: {m.to_class_id}", status.HTTP_400_BAD_REQUEST)

    # Get all ACTIVE records in source year
    result = await db.execute(
        select(StudentAcademicRecord, User.full_name, Section.name)
        .join(User, StudentAcademicRecord.student_id == User.id)
        .join(AcademicYear, StudentAcademicRecord.academic_year_id == AcademicYear.id)
        .join(Section, StudentAcademicRecord.section_id == Section.id)
        .where(
            AcademicYear.tenant_id == tenant_id,
            AcademicYear.id == payload.source_academic_year_id,
            StudentAcademicRecord.status == "ACTIVE",
        )
    )
    rows = result.all()

    promoted_ids: List[UUID] = []
    skipped: List[dict] = []
    actions_list: List[PromotionAction] = []

    for rec, full_name, section_name in rows:
        sid = str(rec.student_id)

        # Check duplicate in target year
        existing = await db.execute(
            select(StudentAcademicRecord).where(
                StudentAcademicRecord.student_id == rec.student_id,
                StudentAcademicRecord.academic_year_id == payload.target_academic_year_id,
            )
        )
        if existing.scalar_one_or_none():
            skipped.append({"student_id": sid, "full_name": full_name, "reason": "Already has record in target academic year"})
            if preview:
                actions_list.append(PromotionAction(student_id=rec.student_id, full_name=full_name, action="SKIPPED", reason="Already in target year"))
            continue

        to_class_id: Optional[UUID] = None
        to_section_id: Optional[UUID] = None
        action_label = "PROMOTED"
        reason_err: Optional[str] = None

        if sid in override_map:
            ov_action, ov_class, ov_section = override_map[sid]
            if ov_action == "RETAIN":
                to_class_id = rec.class_id
                to_section_id = rec.section_id
                action_label = "RETAINED"
            else:  # PROMOTE
                to_class_id = ov_class
                to_section_id = ov_section
                if to_section_id is None:
                    sec_id, err = await _resolve_section_for_promotion(
                        db, tenant_id, to_class_id, "AUTO", None, rec.section_id
                    )
                    if err:
                        reason_err = err
                    else:
                        to_section_id = sec_id
        else:
            # Apply default_class_promotion
            to_class_id = default_class_map.get(str(rec.class_id))
            if to_class_id is None:
                reason_err = f"No default_class_promotion for class {rec.class_id}"
            else:
                sec_id, err = await _resolve_section_for_promotion(
                    db, tenant_id, to_class_id, sec_behavior, section_name, rec.section_id
                )
                if err:
                    reason_err = err
                else:
                    to_section_id = sec_id

        if reason_err:
            skipped.append({"student_id": sid, "full_name": full_name, "reason": reason_err})
            if preview:
                actions_list.append(PromotionAction(student_id=rec.student_id, full_name=full_name, action="SKIPPED", from_class_id=rec.class_id, from_section_id=rec.section_id, reason=reason_err))
            continue

        if to_section_id is None:
            # Fallback: first section of target class
            secs = await section_service.list_sections(db, tenant_id, active_only=True, class_id=to_class_id)
            if secs:
                to_section_id = secs[0].id

        if to_section_id is None:
            skipped.append({"student_id": sid, "full_name": full_name, "reason": "Could not resolve target section"})
            if preview:
                actions_list.append(PromotionAction(student_id=rec.student_id, full_name=full_name, action="SKIPPED", from_class_id=rec.class_id, reason="No target section"))
            continue

        if not preview:
            rec.status = "PROMOTED"
            new_rec = StudentAcademicRecord(
                student_id=rec.student_id,
                academic_year_id=payload.target_academic_year_id,
                class_id=to_class_id,
                section_id=to_section_id,
                roll_number=rec.roll_number,
                status="ACTIVE",
            )
            db.add(new_rec)
        promoted_ids.append(rec.student_id)
        if preview:
            actions_list.append(PromotionAction(
                student_id=rec.student_id,
                full_name=full_name,
                action=action_label,
                from_class_id=rec.class_id,
                from_section_id=rec.section_id,
                to_class_id=to_class_id,
                to_section_id=to_section_id,
            ))

    if preview:
        return StudentBulkPromoteResult(
            promoted_count=len(promoted_ids),
            promoted_ids=promoted_ids,
            skipped=skipped,
            actions=actions_list,
        )
    await db.commit()
    return StudentBulkPromoteResult(
        promoted_count=len(promoted_ids),
        promoted_ids=promoted_ids,
        skipped=skipped,
    )


async def delete_student(db: AsyncSession, tenant_id: UUID, user_id: UUID) -> bool:
    result = await db.execute(
        select(User).where(User.id == user_id, User.tenant_id == tenant_id, User.user_type == "student")
    )
    user = result.scalar_one_or_none()
    if not user:
        return False
    await db.delete(user)
    await db.commit()
    return True
